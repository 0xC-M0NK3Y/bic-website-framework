import openpyxl
from openpyxl_image_loader import SheetImageLoader
import sys
import os
import cv2

def rotate_images(path):
	for filename in os.listdir(path):
		f = os.path.join(path, filename)
		if os.path.isfile(f):
			try:
				img = cv2.imread(f, cv2.IMREAD_UNCHANGED)
				img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
				cv2.imwrite(f, img)
			except:
				print(f'Error rotating {f}')

def main():

	if len(sys.argv) != 3:
		print(f'Usage: {sys.argv[0]} <bic.xlsx> <filename for out csv>')
		exit()

	try:
		pxl_doc = openpyxl.load_workbook(sys.argv[1], data_only=True)
	except:
		print('Bad .xlsx document')
		exit()

	sheet = pxl_doc['achats']
	image_loader = SheetImageLoader(sheet)

	out = open(sys.argv[2], 'w')

	for row in range(2, sheet.max_row):

		# detect end with empty colors cell
		if sheet[row+1][3].value == None:
			print(f'end at row {row+1} : {sheet[row+1][3].value} : {sheet[1][3].value}', file=sys.stderr)
			break

		# save image
		try:
			image = image_loader.get(f'C{row+1}')
			path = f'images/bic_{row-2}.png'
			image.save(path)
		except:
			print(f'error saving image row {row+1}', file=sys.stderr)

		# print csv line
		for col in sheet.iter_cols(1, sheet.max_column):
			if col[0].value == None:
				continue
			if col[0].value == 'image':
				print(path, end='|', file=out)
			else:
				if type(col[row].value) == str:
					print(col[row].value.strip(), end='|', file=out)
				else:
					print(col[row].value, end='|', file=out)
		print('', file=out)
	print('rotating images...')
	rotate_images('images')


if __name__ == '__main__':
	main()